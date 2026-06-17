from flask import Blueprint, render_template, redirect, url_for, flash, request, Response
from flask_login import login_required, current_user
from database.models import Question, Major, QuestionTypeEnum, DifficultyEnum
from database import db
import json
import csv
import io

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False

question_bp = Blueprint('question_bank', __name__)

@question_bp.route("/questions")
@login_required
def list_questions():
    if not current_user.is_teacher():
        flash('无权访问', 'danger')
        return redirect(url_for('home'))

    major_id = request.args.get('major_id')
    q_type = request.args.get('type')

    query = Question.query
    if major_id:
        query = query.filter(Question.major_id == int(major_id))
    if q_type:
        query = query.filter(Question.type == q_type)

    questions = query.all()
    majors = Major.query.all()

    return render_template('questions/list.html', questions=questions, majors=majors,
                          selected_major=major_id, selected_type=q_type)

@question_bp.route("/questions/add", methods=['GET', 'POST'])
@login_required
def add_question():
    if not current_user.is_teacher():
        flash('无权访问', 'danger')
        return redirect(url_for('home'))

    if request.method == 'POST':
        content = request.form['content']
        options = request.form.get('options', '')
        answer = request.form['answer']
        analysis = request.form.get('analysis', '')
        major_id = request.form['major_id']
        q_type = request.form['type']
        difficulty = request.form['difficulty']

        question = Question(
            content=content,
            options=options,
            answer=answer,
            analysis=analysis,
            major_id=major_id,
            type=q_type,
            difficulty=difficulty
        )
        db.session.add(question)
        db.session.commit()

        flash('题目添加成功', 'success')
        return redirect(url_for('question_bank.list_questions'))

    majors = Major.query.all()
    return render_template('questions/add.html', majors=majors,
                          question_types=QuestionTypeEnum.__dict__.keys(),
                          difficulties=DifficultyEnum.__dict__.keys())

@question_bp.route("/questions/edit/<int:id>", methods=['GET', 'POST'])
@login_required
def edit_question(id):
    if not current_user.is_teacher():
        flash('无权访问', 'danger')
        return redirect(url_for('home'))

    question = Question.query.get_or_404(id)

    if request.method == 'POST':
        question.content = request.form['content']
        question.options = request.form.get('options', '')
        question.answer = request.form['answer']
        question.analysis = request.form.get('analysis', '')
        question.major_id = request.form['major_id']
        question.type = request.form['type']
        question.difficulty = request.form['difficulty']

        db.session.commit()
        flash('题目更新成功', 'success')
        return redirect(url_for('question_bank.list_questions'))

    majors = Major.query.all()
    return render_template('questions/edit.html', question=question, majors=majors,
                          question_types=QuestionTypeEnum.__dict__.keys(),
                          difficulties=DifficultyEnum.__dict__.keys())

@question_bp.route("/questions/delete/<int:id>", methods=['POST'])
@login_required
def delete_question(id):
    if not current_user.is_teacher():
        flash('无权访问', 'danger')
        return redirect(url_for('home'))

    question = Question.query.get_or_404(id)
    db.session.delete(question)
    db.session.commit()
    flash('题目删除成功', 'success')
    return redirect(url_for('question_bank.list_questions'))

@question_bp.route("/majors", methods=['GET', 'POST'])
@login_required
def manage_majors():
    if not current_user.is_teacher():
        flash('无权访问', 'danger')
        return redirect(url_for('home'))

    if request.method == 'POST':
        name = request.form['name']
        description = request.form.get('description', '')

        if Major.query.filter_by(name=name).first():
            flash('专业已存在', 'danger')
            return redirect(url_for('question_bank.manage_majors'))

        major = Major(name=name, description=description)
        db.session.add(major)
        db.session.commit()
        flash('专业添加成功', 'success')

    majors = Major.query.all()
    return render_template('questions/majors.html', majors=majors)

@question_bp.route("/majors/delete/<int:id>", methods=['POST'])
@login_required
def delete_major(id):
    if not current_user.is_teacher():
        flash('无权访问', 'danger')
        return redirect(url_for('home'))

    major = Major.query.get_or_404(id)

    # 级联检查：专业下有题目时拒绝删除
    question_count = Question.query.filter_by(major_id=id).count()
    if question_count > 0:
        flash(f'专业「{major.name}」下还有 {question_count} 道题目，请先删除相关题目后再删除专业', 'warning')
        return redirect(url_for('question_bank.manage_majors'))

    db.session.delete(major)
    db.session.commit()
    flash('专业删除成功', 'success')
    return redirect(url_for('question_bank.manage_majors'))


@question_bp.route("/majors/<int:id>/info")
@login_required
def major_info(id):
    """AJAX接口：返回专业下的题目数量"""
    if not current_user.is_teacher():
        return {'error': '无权访问'}, 403
    major = Major.query.get_or_404(id)
    question_count = Question.query.filter_by(major_id=id).count()
    return {'major_id': id, 'name': major.name, 'question_count': question_count}


# ──────────────────────────────────────────────────────────
#  题目导入 / 导出 / 模板下载
# ──────────────────────────────────────────────────────────

_TYPE_LABELS = {
    'single_choice': '单选题', 'multiple_choice': '多选题',
    'fill_blank': '填空题', 'true_false': '判断题',
    'short_answer': '问答题', 'programming': '编程题',
    'application': '应用题', 'calculation': '计算题',
}
_TYPE_REVERSE = {v: k for k, v in _TYPE_LABELS.items()}

_DIFF_LABELS = {'easy': '简单', 'medium': '中等', 'hard': '困难'}
_DIFF_REVERSE = {v: k for k, v in _DIFF_LABELS.items()}

_CSV_HEADERS = ['题目内容', '专业', '题型', '难度', '选项(JSON)', '答案', '解析']


# ─── Excel 工具函数 ────────────────────────────────────────
def _make_excel_workbook(headers, rows=None):
    """创建 Excel 工作簿，返回 bytes"""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = '题目列表'
    # 表头样式
    header_font = Font(name='微软雅黑', bold=True, size=11, color='FFFFFF')
    header_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center', wrap_text=True)
    # 写表头
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align
    # 写数据行
    if rows:
        for r_idx, row in enumerate(rows, 2):
            for c_idx, val in enumerate(row, 1):
                ws.cell(row=r_idx, column=c_idx, value=val)
    # 自动列宽
    col_widths = [40, 16, 12, 10, 40, 20, 40]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[openpyxl.utils.get_column_letter(i)].width = w
    # 冻结表头
    ws.freeze_panes = 'A2'
    buf = io.BytesIO()
    wb.save(buf)
    buf.seek(0)
    return buf.getvalue()


def _read_excel_rows(file_storage):
    """从上传的 Excel 文件中读取行，返回 (headers, data_rows)"""
    wb = openpyxl.load_workbook(file_storage, read_only=True, data_only=True)
    ws = wb.active
    rows_iter = ws.iter_rows(values_only=True)
    try:
        headers = [str(c or '').strip() for c in next(rows_iter)]
    except StopIteration:
        return [], []
    data_rows = []
    for row in rows_iter:
        data_rows.append([str(c or '').strip() if c is not None else '' for c in row])
    return headers, data_rows


@question_bp.route("/questions/template")
@login_required
def download_template():
    """下载题目导入模板（支持 CSV / Excel）"""
    if not current_user.is_teacher():
        flash('无权访问', 'danger')
        return redirect(url_for('home'))

    fmt = request.args.get('format', 'csv').lower()

    if fmt == 'excel':
        if not HAS_OPENPYXL:
            flash('服务器未安装 openpyxl，请使用 CSV 格式或联系管理员安装', 'warning')
            return redirect(url_for('question_bank.list_questions'))
        rows = [[
            'Python中哪个关键字用于定义函数？',
            'Python编程',
            '单选题',
            '简单',
            '["A. class", "B. def", "C. function", "D. define"]',
            'B',
            'def 是 Python 中定义函数的关键字'
        ]]
        data = _make_excel_workbook(_CSV_HEADERS, rows)
        return Response(
            data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=question_template.xlsx'}
        )

    # CSV 默认
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_CSV_HEADERS)
    writer.writerow([
        'Python中哪个关键字用于定义函数？',
        'Python编程',
        '单选题',
        '简单',
        '["A. class", "B. def", "C. function", "D. define"]',
        'B',
        'def 是 Python 中定义函数的关键字'
    ])
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=question_template.csv'}
    )


@question_bp.route("/questions/export")
@login_required
def export_questions():
    """导出当前筛选题目（支持 CSV / Excel）"""
    if not current_user.is_teacher():
        flash('无权访问', 'danger')
        return redirect(url_for('home'))

    major_id = request.args.get('major_id')
    q_type = request.args.get('type')
    fmt = request.args.get('format', 'csv').lower()

    query = Question.query
    if major_id:
        query = query.filter(Question.major_id == int(major_id))
    if q_type:
        query = query.filter(Question.type == q_type)
    questions = query.all()

    rows = []
    for q in questions:
        major_name = q.major.name if q.major else ''
        rows.append([
            q.content,
            major_name,
            _TYPE_LABELS.get(q.type, q.type),
            _DIFF_LABELS.get(q.difficulty, q.difficulty),
            q.options or '',
            q.answer,
            q.analysis or ''
        ])

    if fmt == 'excel':
        if not HAS_OPENPYXL:
            flash('服务器未安装 openpyxl，请使用 CSV 格式', 'warning')
            return redirect(url_for('question_bank.list_questions'))
        data = _make_excel_workbook(_CSV_HEADERS, rows)
        return Response(
            data,
            mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
            headers={'Content-Disposition': 'attachment; filename=questions_export.xlsx'}
        )

    # CSV 默认
    output = io.StringIO()
    writer = csv.writer(output)
    writer.writerow(_CSV_HEADERS)
    writer.writerows(rows)
    return Response(
        '\ufeff' + output.getvalue(),
        mimetype='text/csv',
        headers={'Content-Disposition': 'attachment; filename=questions_export.csv'}
    )


@question_bp.route("/questions/import", methods=['POST'])
@login_required
def import_questions():
    """从 CSV 或 Excel 文件批量导入题目（自动识别文件格式）"""
    if not current_user.is_teacher():
        flash('无权访问', 'danger')
        return redirect(url_for('home'))

    file = request.files.get('file')
    if not file or not file.filename:
        flash('请选择要导入的文件', 'warning')
        return redirect(url_for('question_bank.list_questions'))

    filename = file.filename.lower()
    is_excel = filename.endswith(('.xlsx', '.xls'))

    try:
        if is_excel:
            if not HAS_OPENPYXL:
                flash('服务器未安装 openpyxl，无法解析 Excel 文件，请导出为 CSV 格式后再试', 'warning')
                return redirect(url_for('question_bank.list_questions'))
            headers, data_rows = _read_excel_rows(file)
        else:
            raw = file.read().decode('utf-8-sig')
            reader = csv.DictReader(io.StringIO(raw))
            headers = reader.fieldnames or []
            data_rows = [list(row.values()) for row in reader]

        # 建立列索引映射
        col_map = {h: i for i, h in enumerate(headers)}
        required_cols = ['题目内容', '专业', '题型', '难度', '答案']
        missing_cols = [c for c in required_cols if c not in col_map]
        if missing_cols:
            flash(f'文件表头缺少以下列：{", ".join(missing_cols)}，请确认使用的是正确的模板', 'danger')
            return redirect(url_for('question_bank.list_questions'))

        count = 0
        errors = []
        for i, row in enumerate(data_rows, start=2):
            def get_cell(col_name):
                idx = col_map.get(col_name)
                if idx is None or idx >= len(row):
                    return ''
                return (row[idx] or '').strip()

            content = get_cell('题目内容')
            major_name = get_cell('专业')
            type_label = get_cell('题型')
            diff_label = get_cell('难度')
            options_raw = get_cell('选项(JSON)')
            answer = get_cell('答案')
            analysis = get_cell('解析')

            if not content:
                errors.append(f'第 {i} 行：题目内容为空，已跳过')
                continue
            if not answer:
                errors.append(f'第 {i} 行：答案为空，已跳过')
                continue

            major = Major.query.filter_by(name=major_name).first() if major_name else None
            if not major:
                if major_name:
                    major = Major(name=major_name)
                    db.session.add(major)
                    db.session.flush()
                else:
                    errors.append(f'第 {i} 行：专业为空，已跳过')
                    continue

            q_type = _TYPE_REVERSE.get(type_label, type_label)
            if not q_type:
                q_type = 'single_choice'

            difficulty = _DIFF_REVERSE.get(diff_label, diff_label)
            if difficulty not in ('easy', 'medium', 'hard'):
                difficulty = 'medium'

            question = Question(
                content=content,
                options=options_raw if options_raw else None,
                answer=answer,
                analysis=analysis if analysis else None,
                major_id=major.id,
                type=q_type,
                difficulty=difficulty
            )
            db.session.add(question)
            count += 1

        db.session.commit()

        if errors:
            flash(f'成功导入 {count} 道题目，但有 {len(errors)} 行存在问题：{errors[0]}', 'warning')
        else:
            flash(f'成功导入 {count} 道题目！', 'success')

    except Exception as e:
        db.session.rollback()
        flash(f'导入失败：{str(e)[:100]}', 'danger')

    return redirect(url_for('question_bank.list_questions'))
